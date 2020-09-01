import re

import openpyxl
import pymysql
import datetime

data_dailyfx = {}


class dailyfx:

    def read_dailyfx_sql(self):
        host = '47.102.131.177'
        port = 3306
        user = 'calendar'
        password = 'calendar_2019'
        db = 'economic_calendar'

        the_v = '1'
        start_time = '2019-01-01'
        end_time = '2020-06-20'
        conn = pymysql.Connect(
            host=host,
            port=port,
            user=user,
            password=password,
            db=db)
        conn.ping(reconnect=True)
        cursor = conn.cursor()

        # 取国家和重要性
        select_sql = """SELECT real_date, theme, result, country, importance FROM event_calendar WHERE the_version=%s AND real_date between %s and %s"""
        cursor.execute(select_sql, (the_v, datetime.datetime.strptime(start_time, '%Y-%m-%d').date(),
                                    datetime.datetime.strptime(end_time, '%Y-%m-%d').date()))
        result = cursor.fetchall()
        return result

    def check_theme(self, result):
        for i in result:
            real_date, theme, result, country, importance = i
            real_date = str(real_date)
            if result == '':
                result = '空'
            # 剔除讲话和新闻事件
            # contry_re = re.findall('元', country)
            # contry_re1 = re.findall('元区', country)
            # if contry_re == [] or (contry_re != [] and contry_re1 != []):
            #     theme_re = re.findall('话', theme)
            # theme_news = re.findall('新闻', theme)
            # theme_cen = re.findall('会议', theme)
            # theme_cen = re.findall('休市', theme)
            # if theme_re == []:
            theme_re4 = re.findall('[（](.*?)[）]', theme)
            if theme_re4 != []:
                theme = theme.replace(f'（{theme_re4[0]}）', '')
            theme_re1 = re.findall('[\d+]+月', theme)
            if theme_re1 != []:
                theme = theme.replace(theme_re1[0], '')
            theme_re2 = re.findall('[\d+]+日', theme)
            if theme_re2 != []:
                theme = theme.replace(theme_re2[0], '')
            theme_re3 = re.findall('截至', theme)
            if theme_re3 != []:
                theme = theme.replace(theme_re3[0], '')
            theme_re5 = re.findall('[第](.*?)[季度]', theme)
            if theme_re5 != []:
                theme = theme.replace('第%s季度' % theme_re5[0], '')
            theme_re6 = re.findall('[M][\d+]', theme)
            if theme_re6 != []:
                theme = theme.replace(theme_re6[0], '')
            theme_re7 = re.findall('当周', theme)
            if theme_re7 != []:
                theme = theme.replace(theme_re7[0], '')
            else:
                pass
            # print(theme_re1, theme_re2, theme_re3, theme_re4, theme_re5, theme)
            if theme in data_dailyfx.keys():
                data_dailyfx['%s' % theme].append({real_date: result})
            else:
                data_dailyfx['%s' % theme] = []

    def reduce_num(self):
        # 判断长度，去掉result少的事件
        for k, v in list(data_dailyfx.items()):
            if len(v) > 7:
                data_dailyfx.pop(k)
        print(len(data_dailyfx.keys()))

    def check_null(self):
        # 检查异常时间
        for k, v in data_dailyfx.items():
            for i in range(0, len(v) - 2):
                S_time = list(v[i].keys())[0].split('-')
                E_time = list(v[i + 1].keys())[0].split('-')
                s_time = (int(S_time[0]), int(S_time[1]), int(S_time[2]), 0, 0, 0, 0, 0, 0)
                e_time = (int(E_time[0]), int(E_time[1]), int(E_time[2]), 0, 0, 0, 0, 0, 0)
                import time
                day_num = int(abs(time.mktime(s_time) - time.mktime(e_time)) // 24 // 3600)
                if re.findall('周', k) != [] and day_num > 10:
                    data_dailyfx[k][i] = str(v[i]) + '异常'
                elif day_num > 40:
                    data_dailyfx[k][i] = str(v[i]) + '异常'

                if k == '美国平均每周工时' or k == '美国实际平均周薪年率' and day_num > 40:
                    data_dailyfx[k][i] = str(v[i]) + '异常'
                elif k == '瑞士央行利率决议' or k == '瑞士活期存款利率' or k == '瑞士SECO消费者信心指数' and day_num > 150:
                    data_dailyfx[k][i] = str(v[i]) + '异常'

import xlrd
from xlutils.copy import copy

class excel_save:

    def save_excel(self):
        import copy
        for k, v in data_dailyfx.items():
            list_ = []
            list_.append(k)
            list_.append('次数：%s' % (len(v) - 2))
            for i in v:
                list_.append(str(i))
            # print(list_)
            # write_excel_xlsx('../dailyfx.xls', 'Sheet1', copy.deepcopy(list_))
            list_.clear()

    def read_excel(self, book_name_xlsx):
        from openpyxl import load_workbook
        wb = load_workbook(book_name_xlsx)
        ws = wb.get_sheet_by_name("sheet")
        data_list = {}
        for row in ws.rows:
            num = 1
            for cell in row:
                if num == 1:
                    data_list[f'{cell.value}'] = []
                    key = cell.value
                    num += 1
                else:
                    data_list[f'{key}'].append(cell.value)
            #     print(cell.value, "\t", end="")
            # print()
        return data_list

    def write_excel_xlsx(self, book_name_xlsx, sheet_name_xlsx, value):
        wt = xlrd.open_workbook(book_name_xlsx, formatting_info=True)
        # 复制文件进行操作，直接操作原文件会覆盖
        new_wt = copy(wt)
        new_sheet = new_wt.get_sheet(0)
        table = wt.sheet_by_name(sheet_name_xlsx)
        rows_num = table.nrows
        for num in range(len(value)):
            new_sheet.write(rows_num, num, value[num])
        new_wt.save(book_name_xlsx)

    def write_xlsx(self, path, sheet_name, value):
        index = len(value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        workbook.save(path)

data_investing = {}

class investing:
    def read_investing_sql(self):
        host = '47.100.162.232'
        port = 3306
        user = 'root'
        password = '1qaz@WSX'
        db = 'crawl_spider_information'

        start_time = '2017-06-19'
        end_time = '2020-06-20'

        conn = pymysql.Connect(
            host=host,
            port=port,
            user=user,
            password=password,
            db=db)
        conn.ping(reconnect=True)
        cursor = conn.cursor()

        # 取国家和重要性
        select_sql = """SELECT real_date, real_time, country, theme, importance, result, e_Forecast, b_value, unit FROM investing WHERE real_date between %s and %s"""
        # select_sql = """SELECT real_date, real_time, country, theme, importance, result, e_Forecast, b_value, unit FROM investing WHERE theme=%s and real_date between %s and %s"""
        cursor.execute(select_sql, (datetime.datetime.strptime(start_time, '%Y-%m-%d').date(),
                                    datetime.datetime.strptime(end_time, '%Y-%m-%d').date()))
        result = cursor.fetchall()
        return result

    def check_theme(self, result):
        for i in result:
            real_date, real_time, country, theme, importance, result, e_Forecast, b_value, unit = i
            # country = self.check_country(country)
            real_date = str(real_date).replace('/', '-')
            # if result == '':
            #     result = '空'
            # else:
            #     # 处理数据，取两位数
            #     result = result.replace('.', '').replace('-', '')[0:2]
            # if theme in data_investing.keys():
            #     data_investing['%s' % theme].append({real_date: result})
            # else:
            #     data_investing['%s' % theme] = []

    def check_country(self, country):
        # '欧元区', '德国', '法国', '意大利', 'AUD', 'CAD', 'CNY', 'JPY', 'GBP', 'NZD', 'CHF', 'HKD', 'USD'
        if country == 'AUD':
            country = '澳大利亚'
        elif country == 'CAD':
            country = '加拿大'
        elif country == 'CNY':
            country = '中国'
        elif country == 'JPY':
            country = '日本'
        elif country == 'GBP':
            country = '英国'
        elif country == 'NZD':
            country = '新西兰'
        elif country == 'CHF':
            country = '瑞士'
        elif country == 'HKD':
            country = '香港'
        elif country == 'USD':
            country = '美国'
        return country

def main_investing():
    class_one = investing()
    data_investing_sql = class_one.read_investing_sql()
    # print(data_investing)
    return data_investing

def main_daliyfx():
    class_two = dailyfx()
    data_dailyfx_sql = class_two.read_dailyfx_sql()
    class_two.check_theme(data_dailyfx_sql)
    print(data_dailyfx)
    # '2020年FOMC票委、费城联储主席哈克发表讲话': ['美元', {'2020-01-17': '空'}]

def investing_check():
    data_num = {'num': 0}
    # 判断列表
    list_num = []
    # 保存excel列表
    list_save = []
    res_exc = excel_save().read_excel('../111.xlsx')
    res_inv = main_investing()
    # print(res_exc)
    # print(res_inv)
    print(len(res_exc.keys()), len(res_inv.keys()))
    x = 1
    for exc_key, exc_value in res_exc.items():
        print(x)
        x += 1
        for inv_key, inv_value in res_inv.items():
            if exc_key == inv_key:
                list_save.append([exc_key, inv_key])
                break
            else:
                for i_exc in exc_value:
                    if i_exc == None:
                        break
                    for i_inv in inv_value:
                        list_num.append(i_inv)
                    # 处理excle表中数据key：value，添加到判断列表
                    key1 = list(eval(i_exc).keys())[0]
                    value1 = list(eval(i_exc).values())[0].replace('.', '').replace('-', '')[0:2]
                    # 判断英为数据是否在判断列表里
                    if {key1: value1} in list_num and exc_key[0:2] == inv_key[0:2]:
                        # print(exc_key, inv_key)
                        # print(data_num['num'])
                        # print({key1: value1}, list_num)
                        data_num['num'] += 1
                        list_num.clear()
                    # 事件有10个及以上的值相似，且国家相同，则把两个事件加到一起
                if data_num['num'] >= 5:
                    # 已经添加的pass
                    if [exc_key, inv_key] in list_save:
                        pass
                    else:
                        list_save.append([exc_key, inv_key])
                    data_num['num'] = 0
                    break
                data_num['num'] = 0
                # for i_inv in inv_value:
                #     for i_exc in exc_value:
                #         if i_exc == None:
                #             break
                #         # 处理excle表中数据key：value，添加到判断列表
                #         key1 = list(eval(i_exc).keys())[0]
                #         value1 = list(eval(i_exc).values())[0].replace('.', '').replace('-', '')[0:2]
                #         list_num.append({key1: value1})
                #
                #     # 判断英为数据是否在判断列表里
                #     if i_inv in list_num and exc_key[0:2] == inv_key[0:2]:
                #         # print(exc_key, inv_key)
                #         # print(data_num['num'])
                #         # print(i_inv, list_num)
                #         data_num['num'] += 1
                #         list_num.clear()
                # # 事件有5个及以上的值相似，且国家相同，则把两个事件加到一起
                # if data_num['num'] >= 5:
                #     # 已经添加的pass
                #     if [exc_key, inv_key] in list_save:
                #         pass
                #     else:
                #         list_save.append([exc_key, inv_key])
                #     data_num['num'] = 0
                #     break
                # data_num['num'] = 0
    print(list_save)
    excel_save().write_xlsx('../investing.xlsx', 'Sheet1', list_save)

if __name__ == '__main__':
    host = '47.100.162.232'
    port = 3306
    user = 'root'
    password = '1qaz@WSX'
    db = 'crawl_spider_information'
    conn = pymysql.Connect(
        host=host,
        port=port,
        user=user,
        password=password,
        db=db)

    cursor = conn.cursor()
    # # company_send_epx
    # select_sql = """select theme from company_send_epx where company=%s """

    select_sql = """SELECT theme FROM investing WHERE theme LIKE '%Investing.com%'"""
    cursor.execute(select_sql)
    result = cursor.fetchall()
    conn.close()
    print(result)
    for theme in result:
        print(theme[0].replace('Investing.com', ''))
        new_theme = theme[0].replace('Investing.com', '')
        # select_sql = """insert into company_send_epx (company, theme) values("%s", "%s")""" % ('dingding', theme[0])
        select_sql = """update Investing set theme=%s where theme=%s"""
        try:
            # cursor.execute(select_sql)
            conn.ping(reconnect=True)
            cursor.execute(select_sql, (new_theme, theme[0]))
            conn.commit()
        except Exception as e:
            print(e)
            conn.rollback()